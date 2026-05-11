
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.contrib import messages
from django.shortcuts import render, redirect

# Import đầy đủ Models (Dùng ..models nếu file này nằm trong thư mục views/admin/)
from ...models import (
    SanPham, LoaiSanPham, KhachHang, HoaDon, ChiNhanh, 
    NhanVien, TaiKhoan, TonKhoChiNhanh
)
def import_san_pham_excel(request):
    if request.method == 'POST' and request.FILES.get('file_excel'):
        file = request.FILES['file_excel']
        
        try:
            # Đọc file excel
            df = pd.read_excel(file)
            
            # Chạy vòng lặp từng dòng trong file
            for index, row in df.iterrows():
                # Lấy hoặc tạo Loại sản phẩm (để tránh lỗi khóa ngoại)
                loai_ten = str(row.get('ten_loai', 'Khác'))
                loai_obj, _ = LoaiSanPham.objects.get_or_create(ten_loai=loai_ten)
                
                # Cập nhật nếu trùng tên hoặc tạo mới
                SanPham.objects.update_or_create(
                    ten_san_pham=row['ten_san_pham'],
                    defaults={
                        'loai': loai_obj,
                        'don_gia': row.get('don_gia', 0),
                        'so_luong': row.get('so_luong', 0),
                        'mo_ta': row.get('mo_ta', ''),
                        'gia_hien_tai': row.get('don_gia', 0), # Mặc định bằng đơn giá
                    }
                )
            messages.success(request, f"Đã nhập thành công dữ liệu từ file!")
        except Exception as e:
            messages.error(request, f"Lỗi khi đọc file: {e}")
            
    return redirect('store:admin_product_list')
def admin_dashboard(request):
    tai_khoan_id = request.session.get("tai_khoan_id")
    if not tai_khoan_id:
        return redirect("store:login")

    try:
        tai_khoan = TaiKhoan.objects.get(id=tai_khoan_id)
    except TaiKhoan.DoesNotExist:
        return redirect("store:login")

    is_staff_role = tai_khoan.role == "STAFF"

    nhan_vien_hien_tai = None
    chi_nhanh_hien_tai = None

    if is_staff_role:
        nhan_vien_hien_tai = (
            NhanVien.objects
            .select_related("chi_nhanh")
            .filter(tai_khoan=tai_khoan)
            .first()
        )
        if nhan_vien_hien_tai:
            chi_nhanh_hien_tai = nhan_vien_hien_tai.chi_nhanh

    if is_staff_role and chi_nhanh_hien_tai:
        # ── STAFF: chỉ xem dữ liệu chi nhánh của mình ──────────────

        # HoaDon liên kết chi nhánh qua nhan_vien__chi_nhanh
        don_hang_qs = HoaDon.objects.filter(
            nhan_vien__chi_nhanh=chi_nhanh_hien_tai
        )

        don_hang_moi = (
            don_hang_qs
            .select_related("khach_hang")
            .order_by("-ngay_lap")[:10]
        )

        # Sản phẩm sắp hết: dùng so_luong_ton <= muc_can_canh_bao (field động)
        san_pham_sap_het = (
            TonKhoChiNhanh.objects
            .select_related("san_pham", "san_pham__loai")
            .filter(
                chi_nhanh=chi_nhanh_hien_tai,
                so_luong_ton__gt=0,
            )
            .extra(where=["so_luong_ton <= muc_can_canh_bao"])  # dùng muc_can_canh_bao động
            .order_by("so_luong_ton")[:10]
        )

        context = {
            "is_staff_role": True,
            "nhan_vien_hien_tai": nhan_vien_hien_tai,
            "chi_nhanh_hien_tai": chi_nhanh_hien_tai,
            "tong_san_pham": TonKhoChiNhanh.objects.filter(
                chi_nhanh=chi_nhanh_hien_tai
            ).count(),
            "tong_khach_hang": KhachHang.objects.count(),
            "tong_don_hang": don_hang_qs.count(),
            "don_hang_moi": don_hang_moi,
            "san_pham_sap_het": san_pham_sap_het,
        }

    else:
        # ── ADMIN: xem toàn bộ hệ thống ────────────────────────────

        don_hang_moi = (
            HoaDon.objects
            .select_related("khach_hang")
            .order_by("-ngay_lap")[:10]
        )

        # Admin dùng SanPham.so_luong (tổng toàn hệ thống)
        san_pham_sap_het = (
            SanPham.objects
            .select_related("loai")
            .filter(so_luong__gt=0, so_luong__lte=5)
            .order_by("so_luong")[:10]
        )

        context = {
            "is_staff_role": False,
            "nhan_vien_hien_tai": None,
            "chi_nhanh_hien_tai": None,
            "tong_san_pham": SanPham.objects.count(),
            "tong_khach_hang": KhachHang.objects.count(),
            "tong_don_hang": HoaDon.objects.count(),
            "tong_chi_nhanh": ChiNhanh.objects.count(),
            "tong_nhan_vien": NhanVien.objects.count(),
            "tong_tai_khoan": TaiKhoan.objects.count(),
            "don_hang_moi": don_hang_moi,
            "san_pham_sap_het": san_pham_sap_het,
        }

    return render(request, "store/admin/dashboard.html", context)
def export_dashboard_excel(request):
    # 1. KIỂM TRA QUYỀN (Phân biệt Admin và Staff)
    tai_khoan_id = request.session.get("tai_khoan_id")
    if not tai_khoan_id:
        return HttpResponse("Bạn cần đăng nhập", status=401)
    
    tai_khoan = TaiKhoan.objects.get(id=tai_khoan_id)
    is_staff_role = tai_khoan.role == "STAFF"
    chi_nhanh_hien_tai = None

    if is_staff_role:
        nv = NhanVien.objects.filter(tai_khoan=tai_khoan).select_related("chi_nhanh").first()
        if nv:
            chi_nhanh_hien_tai = nv.chi_nhanh
    # 👇 THÊM ĐOẠN LẤY THÁNG/NĂM NÀY VÀO 👇
    thang_nam = request.GET.get('thang_nam') # Định dạng nhận được: "2026-05"
    # 2. KHỞI TẠO FILE EXCEL
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="008A37", end_color="008A37", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # SHEET 1: BÁO CÁO DOANH THU
    # ==========================================
    ws1 = wb.active
    ws1.title = "Báo cáo Doanh Thu"
    
    headers_dt = ['Mã Đơn', 'Ngày Lập', 'Khách Hàng', 'Chi Nhánh', 'Thanh Toán', 'Tổng Tiền (VNĐ)', 'Trạng Thái']
    ws1.append(headers_dt)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Lọc dữ liệu theo quyền: Staff chỉ thấy hóa đơn của chi nhánh mình
    orders_qs = HoaDon.objects.exclude(trang_thai__in=['CHO_XU_LY', 'HUY'])
    if is_staff_role and chi_nhanh_hien_tai:
        orders_qs = orders_qs.filter(nhan_vien__chi_nhanh=chi_nhanh_hien_tai)
    # 👇 THÊM ĐIỀU KIỆN LỌC THEO THÁNG 👇
    thong_tin_thang = "Tất cả"
    if thang_nam:
        nam, thang = thang_nam.split('-')
        orders_qs = orders_qs.filter(ngay_lap__year=nam, ngay_lap__month=thang)
        thong_tin_thang = f"Tháng {thang}/{nam}"
    # 👆 ============================= 👆
    orders = orders_qs.select_related('khach_hang', 'nhan_vien__chi_nhanh').prefetch_related('chi_tiets')
    
    tong_tat_ca = 0
    for order in orders:
        tong_tien_don = sum(item.so_luong * item.don_gia for item in order.chi_tiets.all())
        tong_tat_ca += tong_tien_don
        
        # Dịch phương thức thanh toán
        pt_thanh_toan = order.phuong_thuc_thanh_toan
        if pt_thanh_toan in ['CASH', 'TIEN_MAT']:
            ten_thanh_toan = "Tiền mặt"
        elif pt_thanh_toan in ['BANK', 'CHUYEN_KHOAN']:
            ten_thanh_toan = "Chuyển khoản"
        else:
            ten_thanh_toan = str(pt_thanh_toan) if pt_thanh_toan else "Chưa xác định"
            
        ws1.append([
            order.id,
            localtime(order.ngay_lap).strftime("%d/%m/%Y %H:%M"),
            order.khach_hang.ten_khach_hang if order.khach_hang else "Khách lẻ",
            order.nhan_vien.chi_nhanh.ten_chi_nhanh if order.nhan_vien and order.nhan_vien.chi_nhanh else "Online",
            ten_thanh_toan,
            tong_tien_don,
            order.get_trang_thai_display()
        ])
    
    # Dòng tổng kết
    last_row = ws1.max_row + 1
    ws1.append(['', '', '', '', 'TỔNG DOANH THU:', tong_tat_ca, ''])
    ws1[f'F{last_row}'].number_format = '#,##0"đ"'
    ws1[f'E{last_row}'].font = Font(bold=True)
    ws1[f'F{last_row}'].font = Font(bold=True, color="D8232A")

    # ==========================================
    # SHEET 2: BÁO CÁO TỒN KHO
    # ==========================================
    ws2 = wb.create_sheet("Báo cáo Tồn Kho")
    headers_kho = ['Chi Nhánh', 'Tên Sản Phẩm', 'Loại', 'Số Lượng Tồn', 'Mức Cảnh Báo', 'Trạng Thái']
    ws2.append(headers_kho)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Lọc tồn kho theo quyền: Staff chỉ xem kho của mình
    ton_kho_qs = TonKhoChiNhanh.objects.select_related('chi_nhanh', 'san_pham', 'san_pham__loai')
    if is_staff_role and chi_nhanh_hien_tai:
        ton_kho_qs = ton_kho_qs.filter(chi_nhanh=chi_nhanh_hien_tai)

    for item in ton_kho_qs:
        ws2.append([
            item.chi_nhanh.ten_chi_nhanh,
            item.san_pham.ten_san_pham,
            item.san_pham.loai.ten_loai,
            item.so_luong_ton,
            item.muc_can_canh_bao,
            "SẮP HẾT HÀNG" if item.so_luong_ton <= item.muc_can_canh_bao else "Ổn định"
        ])

    # Tự động chỉnh độ rộng cột cho đẹp
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 22

    # 3. XUẤT FILE VỚI TÊN ĐỘNG
    # 👇 ĐỔI TÊN FILE TẢI VỀ CÓ KÈM THÁNG 👇
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    prefix = "ChiNhanh" if is_staff_role else "HeThong"
    file_name = f'Bao_Cao_{prefix}_{thong_tin_thang.replace("/", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    wb.save(response)

    return response