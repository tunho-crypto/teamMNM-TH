import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from .models import HoaDon, TonKhoChiNhanh

def export_revenue_and_stock_excel():
    # 1. Khởi tạo Workbook
    wb = openpyxl.Workbook()
    
    # --- SHEET 1: BÁO CÁO DOANH THU ---
    ws1 = wb.active
    ws1.title = "Báo cáo Doanh Thu"
    
    # Định dạng Header
    header_fill = PatternFill(start_color="008A37", end_color="008A37", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['Mã Đơn', 'Ngày Lập', 'Khách Hàng', 'Chi Nhánh', 'Thanh Toán', 'Tổng Tiền (VNĐ)']
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Lấy dữ liệu Hóa đơn
    orders = HoaDon.objects.filter(trang_thai='HOAN_TAT').select_related('khach_hang', 'nhan_vien__chi_nhanh')
    for order in orders:
        # Tính tổng tiền từ chi tiết hóa đơn
        total = sum(item.so_luong * item.don_gia for item in order.chi_tiets.all())
        ws1.append([
            order.id,
            order.ngay_lap.strftime("%d/%m/%Y %H:%M"),
            order.khach_hang.ten_khach_hang if order.khach_hang else "Khách lẻ",
            order.nhan_vien.chi_nhanh.ten_chi_nhanh if order.nhan_vien and order.nhan_vien.chi_nhanh else "Online",
            order.get_phuong_thuc_thanh_toan_display(),
            total
        ])

    # --- SHEET 2: BÁO CÁO TỒN KHO CHI NHÁNH ---
    ws2 = wb.create_sheet("Tồn Kho Chi Nhánh")
    ws2.append(['Chi Nhánh', 'Sản Phẩm', 'Số Lượng Tồn', 'Trạng Thái'])
    
    # Lấy dữ liệu tồn kho
    stocks = TonKhoChiNhanh.objects.select_related('chi_nhanh', 'san_pham').all()
    for s in stocks:
        ws2.append([
            s.chi_nhanh.ten_chi_nhanh,
            s.san_pham.ten_san_pham,
            s.so_luong_ton,
            "Sắp hết" if s.so_luong_ton <= s.muc_can_canh_bao else "Ổn định"
        ])

    return wb