import os
import sys
import django
import shutil
import random
from decimal import Decimal
from datetime import datetime, timedelta

from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password

 
# 1. DJANGO SETUP
 
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "quanly_bhx_core.settings")
django.setup()

from django.conf import settings
from django.utils import timezone
from store.models import (
    ChiNhanh,
    ChiNhanhImage,
    NhanVien,
    KhachHang,
    LoaiSanPham,
    SanPham,
    SanPhamImage,
    NhaCungCap,
    NhapHang,
    ChiTietNhapHang,
    HoaDon,
    ChiTietHoaDon,
    TaiKhoan,
    TonKhoChiNhanh,
)

 
# 2. FALLBACK DATA
 
ADDRESSES = [
    "12 Nguyễn Trãi, Quận 1, TP.HCM",
    "85 Lê Văn Sỹ, Quận 3, TP.HCM",
    "42 Phan Xích Long, Phú Nhuận, TP.HCM",
    "219 Cách Mạng Tháng 8, Quận 10, TP.HCM",
    "88 Nguyễn Oanh, Gò Vấp, TP.HCM",
    "156 Kha Vạn Cân, Thủ Đức, TP.HCM",
    "29 Tô Ký, Quận 12, TP.HCM",
    "61 Hậu Giang, Quận 6, TP.HCM",
    "73 Bình Long, Bình Tân, TP.HCM",
    "104 Lê Trọng Tấn, Tân Phú, TP.HCM",
]

PAYMENTS = ["TIEN_MAT", "CHUYEN_KHOAN"]


 
# 3. FILE PATHS
 
def get_excel_path():
    return os.path.join(BASE_DIR, "data", "bhx_seed_data_full_chuan.xlsx")


 
# 4. HELPER FUNCTIONS
 
def safe_str(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def safe_int(value, default=0):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except Exception:
        try:
            return int(float(str(value).replace(",", "")))
        except Exception:
            return default


def safe_decimal(value, default="0"):
    if value in (None, ""):
        return Decimal(str(default))
    try:
        s = str(value).replace(",", "").strip()
        return Decimal(s)
    except Exception:
        return Decimal(str(default))


def safe_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ["1", "true", "yes", "y", "x", "có"]


def normalize_text(value):
    return safe_str(value).strip().lower()


def build_header_map(ws):
    headers = []
    for cell in ws[1]:
        headers.append(normalize_text(cell.value))
    return {header: idx for idx, header in enumerate(headers) if header}


def get_by_header(row, header_map, *names, default=None):
    for name in names:
        idx = header_map.get(normalize_text(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return default


def copy_image_to_media(ten_file, thu_muc_dich):
    ten_file = safe_str(ten_file)
    if not ten_file:
        return None

    static_dir = os.path.join(settings.BASE_DIR, "static", "images")
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    src = os.path.join(static_dir, ten_file)
    dst_rel = f"{thu_muc_dich}/{ten_file}"
    dst_abs = os.path.join(settings.MEDIA_ROOT, dst_rel)
    os.makedirs(os.path.dirname(dst_abs), exist_ok=True)

    if os.path.exists(src):
        shutil.copy2(src, dst_abs)
        return dst_rel

    print(f"  ⚠️ Không tìm thấy ảnh: {src}")
    return None


def tao_anh_cho_san_pham(san_pham, danh_sach_anh):
    for index, ten_file in enumerate(danh_sach_anh):
        path = copy_image_to_media(ten_file, "san_pham")
        if path:
            SanPhamImage.objects.create(
                san_pham=san_pham,
                image=path,
                la_anh_chinh=(index == 0),
                thu_tu=index,
            )


def tao_anh_cho_chi_nhanh(chi_nhanh, danh_sach_anh):
    if not danh_sach_anh:
        return 0

    ChiNhanhImage.objects.filter(chi_nhanh=chi_nhanh).delete()
    tong = 0

    for index, ten_file in enumerate(danh_sach_anh):
        path = copy_image_to_media(ten_file, "store_gallery")
        if not path:
            continue

        if index == 0:
            chi_nhanh.hinh_anh = path
            chi_nhanh.save(update_fields=["hinh_anh"])

        ChiNhanhImage.objects.create(chi_nhanh=chi_nhanh, image=path)
        tong += 1

    return tong


def create_tai_khoan(email, ho_ten, dien_thoai, role, password, trang_thai="ACTIVE"):
    email = safe_str(email, None)
    dien_thoai = safe_str(dien_thoai, None)
    ho_ten = safe_str(ho_ten)
    role = safe_str(role, "USER")
    trang_thai = safe_str(trang_thai, "ACTIVE")

    tai_khoan, created = TaiKhoan.objects.get_or_create(
        email=email,
        defaults={
            "password_hash": make_password(password),
            "ho_ten": ho_ten,
            "dien_thoai": dien_thoai,
            "role": role,
            "trang_thai": trang_thai,
        },
    )

    if not created:
        changed = False
        if tai_khoan.ho_ten != ho_ten:
            tai_khoan.ho_ten = ho_ten
            changed = True
        if tai_khoan.dien_thoai != dien_thoai:
            tai_khoan.dien_thoai = dien_thoai
            changed = True
        if tai_khoan.role != role:
            tai_khoan.role = role
            changed = True
        if tai_khoan.trang_thai != trang_thai:
            tai_khoan.trang_thai = trang_thai
            changed = True
        if changed:
            tai_khoan.save()

    return tai_khoan


def upsert_nhan_vien_from_tai_khoan(tai_khoan, ten_nhan_vien, chuc_vu, luong, chi_nhanh):
    nhan_vien = NhanVien.objects.filter(tai_khoan=tai_khoan).first()

    if nhan_vien:
        nhan_vien.ten_nhan_vien = safe_str(ten_nhan_vien)
        nhan_vien.chuc_vu = safe_str(chuc_vu)
        nhan_vien.luong = safe_decimal(luong)
        nhan_vien.chi_nhanh = chi_nhanh
        nhan_vien.save()
        return nhan_vien

    return NhanVien.objects.create(
        tai_khoan=tai_khoan,
        ten_nhan_vien=safe_str(ten_nhan_vien),
        chuc_vu=safe_str(chuc_vu),
        luong=safe_decimal(luong),
        chi_nhanh=chi_nhanh,
    )


def create_admin(chi_nhanh):
    admin_acc = create_tai_khoan(
        email="admin@bhx.vn",
        ho_ten="Quản trị viên",
        dien_thoai="0909000000",
        role="ADMIN",
        password="admin123",
    )

    admin_nv = upsert_nhan_vien_from_tai_khoan(
        tai_khoan=admin_acc,
        ten_nhan_vien="Nguyễn Văn An",
        chuc_vu="Quản lý",
        luong="18000000",
        chi_nhanh=chi_nhanh,
    )
    return admin_acc, admin_nv


def create_staff_account_and_employee(index, chi_nhanh, chuc_vu):
    stt = str(index).zfill(4)
    ho_ten = f"Nhân viên {stt}"
    email = f"staff{stt}@bhx.local"
    dien_thoai = f"09{index:08d}"

    tai_khoan = create_tai_khoan(
        email=email,
        ho_ten=ho_ten,
        dien_thoai=dien_thoai,
        role="STAFF",
        password="123456",
    )

    nhan_vien = upsert_nhan_vien_from_tai_khoan(
        tai_khoan=tai_khoan,
        ten_nhan_vien=ho_ten,
        chuc_vu=chuc_vu,
        luong="2000000",
        chi_nhanh=chi_nhanh,
    )

    return tai_khoan, nhan_vien


def clear_old_data():
    ChiTietHoaDon.objects.all().delete()
    HoaDon.objects.all().delete()
    ChiTietNhapHang.objects.all().delete()
    NhapHang.objects.all().delete()
    NhaCungCap.objects.all().delete()
    TonKhoChiNhanh.objects.all().delete()
    SanPhamImage.objects.all().delete()
    SanPham.objects.all().delete()
    LoaiSanPham.objects.all().delete()
    KhachHang.objects.all().delete()
    NhanVien.objects.all().delete()
    TaiKhoan.objects.all().delete()
    ChiNhanhImage.objects.all().delete()
    ChiNhanh.objects.all().delete()


 
# 5. IMPORT EXCEL FUNCTIONS
 
def import_addresses_from_excel(wb):
    global ADDRESSES
    if "dia_chi_mau" not in wb.sheetnames:
        return ADDRESSES

    ws = wb["dia_chi_mau"]
    header_map = build_header_map(ws)
    ds = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        dia_chi = safe_str(get_by_header(row, header_map, "dia_chi", "địa_chỉ", default=row[1] if len(row) > 1 else row[0]))
        if dia_chi:
            ds.append(dia_chi)

    if ds:
        ADDRESSES = ds
    return ADDRESSES


def import_payments_from_excel(wb):
    global PAYMENTS
    if "thanh_toan" not in wb.sheetnames:
        return PAYMENTS

    ws = wb["thanh_toan"]
    header_map = build_header_map(ws)
    ds = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = safe_str(get_by_header(row, header_map, "phuong_thuc", "thanh_toan", default=row[1] if len(row) > 1 else row[0]))
        if value:
            ds.append(value)

    if ds:
        PAYMENTS = ds
    return PAYMENTS


def import_chi_nhanh_from_excel(wb):
    if "chi_nhanh" not in wb.sheetnames:
        raise RuntimeError("File Excel thiếu sheet 'chi_nhanh'.")

    ws = wb["chi_nhanh"]
    header_map = build_header_map(ws)
    result = []
    tong_anh = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        ten = safe_str(get_by_header(row, header_map, "ten_chi_nhanh"))
        dia_chi = safe_str(get_by_header(row, header_map, "dia_chi"))
        latitude = get_by_header(row, header_map, "latitude")
        longitude = get_by_header(row, header_map, "longitude")
        dien_thoai = safe_str(get_by_header(row, header_map, "dien_thoai"), "19001908") or "19001908"

        if not ten:
            continue

        cn = ChiNhanh.objects.create(
            ten_chi_nhanh=ten,
            dia_chi=dia_chi,
            latitude=latitude,
            longitude=longitude,
            dien_thoai=dien_thoai,
        )
        result.append(cn)

        danh_sach_anh = []
        for key in ["anh_1", "anh_2", "anh_3", "anh_4", "anh_5", "anh_6"]:
            value = safe_str(get_by_header(row, header_map, key))
            if value:
                danh_sach_anh.append(value)

        tong_anh += tao_anh_cho_chi_nhanh(cn, danh_sach_anh)

    print(f"  ✅ Đã import {len(result)} chi nhánh")
    print(f"  ✅ Đã import {tong_anh} ảnh cửa hàng từ sheet chi_nhanh")
    return result


def import_customers_from_excel(wb):
    if "khach_hang" not in wb.sheetnames:
        print("  ℹ️ Không có sheet 'khach_hang', chỉ tạo Khách lẻ.")
        khach_le = KhachHang.objects.create(
            ten_khach_hang="Khách lẻ",
            dien_thoai="0901999999",
            dia_chi="Mua tại cửa hàng",
            tai_khoan=None,
        )
        return [], [], khach_le

    ws = wb["khach_hang"]
    header_map = build_header_map(ws)
    created_accounts = []
    created_customers = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        name = safe_str(get_by_header(row, header_map, "ten_khach_hang", "name", "ho_ten"))
        phone = safe_str(get_by_header(row, header_map, "dien_thoai", "phone"))
        email = safe_str(get_by_header(row, header_map, "email"))
        password = safe_str(get_by_header(row, header_map, "password", default="User@123")) or "User@123"

        if not name or not email:
            continue

        tai_khoan = create_tai_khoan(
            email=email,
            ho_ten=name,
            dien_thoai=phone,
            role="USER",
            password=password,
        )

        khach_hang = KhachHang.objects.create(
            ten_khach_hang=name,
            dien_thoai=phone,
            dia_chi=random.choice(ADDRESSES) if ADDRESSES else "TP.HCM",
            tai_khoan=tai_khoan,
        )

        created_accounts.append(tai_khoan)
        created_customers.append(khach_hang)

    khach_le = KhachHang.objects.create(
        ten_khach_hang="Khách lẻ",
        dien_thoai="0901999999",
        dia_chi="Mua tại cửa hàng",
        tai_khoan=None,
    )
    return created_accounts, created_customers, khach_le


def get_or_create_loai_san_pham(ten_loai, cache):
    ten_loai = safe_str(ten_loai, "Khác")
    if ten_loai in cache:
        return cache[ten_loai]
    loai, _ = LoaiSanPham.objects.get_or_create(ten_loai=ten_loai)
    cache[ten_loai] = loai
    return loai


def import_products_from_excel(wb):
    if "san_pham" not in wb.sheetnames:
        raise RuntimeError("File Excel thiếu sheet 'san_pham'.")

    ws = wb["san_pham"]
    header_map = build_header_map(ws)
    loai_cache = {}
    san_phams = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        ten_san_pham = safe_str(get_by_header(row, header_map, "ten_san_pham"))
        don_gia = get_by_header(row, header_map, "don_gia")
        so_luong = get_by_header(row, header_map, "so_luong_chung", "so_luong")
        ten_loai = safe_str(get_by_header(row, header_map, "ten_loai"), "Khác")
        is_khuyen_mai = get_by_header(row, header_map, "is_khuyen_mai")
        gia_goc = get_by_header(row, header_map, "gia_goc")
        gia_khuyen_mai = get_by_header(row, header_map, "gia_khuyen_mai")
        mo_ta = safe_str(get_by_header(row, header_map, "mo_ta"))

        if not ten_san_pham:
            continue

        danh_sach_anh = []
        for key in ["anh_1", "anh_2", "anh_3", "anh_4", "anh_5"]:
            value = safe_str(get_by_header(row, header_map, key))
            if value:
                danh_sach_anh.append(value)

        loai = get_or_create_loai_san_pham(ten_loai, loai_cache)
        is_km = safe_bool(is_khuyen_mai)

        sp = SanPham.objects.create(
            ten_san_pham=ten_san_pham,
            don_gia=safe_decimal(don_gia),
            so_luong=safe_int(so_luong),
            loai=loai,
            mo_ta=mo_ta,
            is_khuyen_mai=is_km,
            gia_goc=safe_decimal(gia_goc) if gia_goc not in (None, "") else None,
            gia_khuyen_mai=safe_decimal(gia_khuyen_mai) if gia_khuyen_mai not in (None, "") else None,
            thanh_phan=safe_str(get_by_header(row, header_map, "thanh_phan")) or None,
            cong_dung=safe_str(get_by_header(row, header_map, "cong_dung")) or None,
            cach_su_dung=safe_str(get_by_header(row, header_map, "cach_su_dung")) or None,
            bao_quan=safe_str(get_by_header(row, header_map, "bao_quan")) or None,
            quy_cach_dong_goi=safe_str(get_by_header(row, header_map, "quy_cach_dong_goi")) or None,
            xuat_xu=safe_str(get_by_header(row, header_map, "xuat_xu")) or None,
            han_su_dung=safe_str(get_by_header(row, header_map, "han_su_dung")) or None,
            luu_y=safe_str(get_by_header(row, header_map, "luu_y")) or None,
        )

        tao_anh_cho_san_pham(sp, danh_sach_anh)
        san_phams.append(sp)

    print(f"  ✅ Đã import {len(san_phams)} sản phẩm kèm mô tả chi tiết")
    return san_phams


def import_ton_kho_chi_nhanh_from_excel(wb):
    if "ton_kho_chi_nhanh" not in wb.sheetnames:
        print("  ℹ️ Không có sheet 'ton_kho_chi_nhanh', bỏ qua import tồn kho theo chi nhánh.")
        return 0

    ws = wb["ton_kho_chi_nhanh"]
    header_map = build_header_map(ws)

    chi_nhanh_map = {cn.ten_chi_nhanh.strip(): cn for cn in ChiNhanh.objects.all()}
    san_pham_map = {sp.ten_san_pham.strip(): sp for sp in SanPham.objects.all()}
    tong = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        ten_chi_nhanh = safe_str(get_by_header(row, header_map, "ten_chi_nhanh"))
        ten_san_pham = safe_str(get_by_header(row, header_map, "ten_san_pham"))
        so_luong_ton = safe_int(get_by_header(row, header_map, "so_luong_ton"), 0)
        muc_can_canh_bao = safe_int(get_by_header(row, header_map, "muc_can_canh_bao"), 5)
        is_hien_thi = safe_bool(get_by_header(row, header_map, "is_hien_thi", default=True))

        if not ten_chi_nhanh or not ten_san_pham:
            continue

        chi_nhanh = chi_nhanh_map.get(ten_chi_nhanh)
        san_pham = san_pham_map.get(ten_san_pham)

        if not chi_nhanh:
            print(f"  ⚠️ Không tìm thấy chi nhánh: {ten_chi_nhanh}")
            continue
        if not san_pham:
            print(f"  ⚠️ Không tìm thấy sản phẩm: {ten_san_pham}")
            continue

        TonKhoChiNhanh.objects.update_or_create(
            chi_nhanh=chi_nhanh,
            san_pham=san_pham,
            defaults={
                "so_luong_ton": so_luong_ton,
                "muc_can_canh_bao": muc_can_canh_bao,
                "is_hien_thi": is_hien_thi,
            },
        )
        tong += 1

    return tong


 
# 6. GENERATED DATA
 
def create_suppliers_and_imports(san_phams):
    if not san_phams:
        return

    ncc1 = NhaCungCap.objects.create(
        ten_ncc="Công ty Nông Sản Sạch",
        dien_thoai="02839990001",
        dia_chi="Hóc Môn, TP.HCM",
    )
    ncc2 = NhaCungCap.objects.create(
        ten_ncc="Công ty Thực Phẩm Tươi",
        dien_thoai="02839990002",
        dia_chi="Bình Chánh, TP.HCM",
    )

    nh1 = NhapHang.objects.create(
        ngay_nhap=timezone.make_aware(datetime(2026, 1, 10, 8, 15)),
        nha_cung_cap=ncc1,
    )
    nh2 = NhapHang.objects.create(
        ngay_nhap=timezone.make_aware(datetime(2026, 1, 15, 9, 0)),
        nha_cung_cap=ncc2,
    )

    for sp in san_phams[:10]:
        ChiTietNhapHang.objects.create(
            nhap_hang=nh1,
            san_pham=sp,
            so_luong=200,
            don_gia_nhap=sp.don_gia * Decimal("0.7"),
        )

    for sp in san_phams[10:20]:
        ChiTietNhapHang.objects.create(
            nhap_hang=nh2,
            san_pham=sp,
            so_luong=120,
            don_gia_nhap=sp.don_gia * Decimal("0.75"),
        )


def create_orders(created_customers, khach_le, staffs, san_phams, admin_nv):
    if len(san_phams) < 3:
        print("  ℹ️ Không đủ sản phẩm để tạo hóa đơn mẫu.")
        return

    if not created_customers:
        created_customers = [khach_le]

    staff_candidates = [s for s in staffs if s.chi_nhanh_id] or [admin_nv]

    for i in range(40):
        if i < 30 and created_customers:
            kh = created_customers[i % len(created_customers)]
        else:
            kh = khach_le if i % 5 == 0 else random.choice(created_customers)

        nhan_vien = random.choice(staff_candidates)
        order_time = timezone.now() - timedelta(
            days=random.randint(0, 45),
            hours=random.randint(0, 23),
            minutes=random.randint(0, 59),
        )

        dia_chi = kh.dia_chi or (random.choice(ADDRESSES) if ADDRESSES else "TP.HCM")
        sdt = kh.dien_thoai or "0901999999"
        ten = kh.ten_khach_hang or "Khách lẻ"

        hoa_don = HoaDon.objects.create(
            ngay_lap=order_time,
            khach_hang=kh,
            nhan_vien=nhan_vien,
            dia_chi_giao_hang=dia_chi,
            sdt_nguoi_nhan=sdt,
            ten_nguoi_nhan=ten,
            trang_thai="HOAN_TAT",
            phuong_thuc_thanh_toan=random.choice(PAYMENTS) if PAYMENTS else "TIEN_MAT",
        )

        chosen_products = random.sample(san_phams, k=random.randint(2, min(5, len(san_phams))))
        for sp in chosen_products:
            don_gia = sp.gia_hien_tai if hasattr(sp, "gia_hien_tai") else sp.don_gia
            ChiTietHoaDon.objects.create(
                hoa_don=hoa_don,
                san_pham=sp,
                so_luong=random.randint(1, 4),
                don_gia=Decimal(str(don_gia)),
            )


 
# 7. RUN IMPORT
 
def run():
    global ADDRESSES, PAYMENTS

    print("Đang chạy import data từ Excel...")
    excel_path = get_excel_path()
    if not os.path.exists(excel_path):
        raise RuntimeError(f"Không tìm thấy file Excel: {excel_path}")

    wb = load_workbook(excel_path)
    print("📄 Các sheet hiện có:", wb.sheetnames)

    print("🧹 Đang dọn dẹp dữ liệu cũ...")
    clear_old_data()

    print("🏠 Đang import địa chỉ mẫu...")
    ADDRESSES = import_addresses_from_excel(wb)

    print("💳 Đang import phương thức thanh toán...")
    PAYMENTS = import_payments_from_excel(wb)

    print("📍 Đang import chi nhánh từ Excel...")
    danh_sach_chi_nhanh_db = import_chi_nhanh_from_excel(wb)
    if not danh_sach_chi_nhanh_db:
        raise RuntimeError("Không có chi nhánh để seed dữ liệu.")

    print("👑 Đang tạo tài khoản admin...")
    admin_acc, admin_nv = create_admin(danh_sach_chi_nhanh_db[0])

    print("👨‍💼 Đang tạo 2 nhân viên cho mỗi chi nhánh...")
    danh_sach_staff_nv = []
    staff_index = 1
    for chi_nhanh in danh_sach_chi_nhanh_db:
        _, nv_thu_ngan = create_staff_account_and_employee(staff_index, chi_nhanh, "Thu ngân")
        staff_index += 1
        _, nv_kho = create_staff_account_and_employee(staff_index, chi_nhanh, "Nhân viên kho")
        staff_index += 1
        danh_sach_staff_nv.extend([nv_thu_ngan, nv_kho])
    print(f"  ✅ Đã tạo {len(danh_sach_staff_nv)} nhân viên staff")

    print("👥 Đang import khách hàng từ Excel...")
    created_accounts, created_customers, khach_le = import_customers_from_excel(wb)

    print("📦 Đang import sản phẩm từ Excel...")
    san_phams = import_products_from_excel(wb)

    print("🏬 Đang import tồn kho theo từng chi nhánh...")
    tong_ton_kho = import_ton_kho_chi_nhanh_from_excel(wb)
    print(f"  ✅ Đã import {tong_ton_kho} dòng tồn kho chi nhánh")
    print(f"📚 Tồn kho chi nhánh: {TonKhoChiNhanh.objects.count()}")

    print("🚚 Đang tạo nhà cung cấp và nhập hàng...")
    create_suppliers_and_imports(san_phams)

    print("🧾 Đang tạo 40 hóa đơn hoàn tất...")
    create_orders(
        created_customers=created_customers,
        khach_le=khach_le,
        staffs=danh_sach_staff_nv,
        san_phams=san_phams,
        admin_nv=admin_nv,
    )

    print("\n🎉 Import từ Excel THÀNH CÔNG!")
    print(f"📍 Chi nhánh: {ChiNhanh.objects.count()}")
    print(f"🖼️ Ảnh chi nhánh: {ChiNhanhImage.objects.count()}")
    print(f"👨‍💼 Staff: {NhanVien.objects.filter(tai_khoan__role='STAFF').count()}")
    print(f"👤 Khách hàng: {KhachHang.objects.count()}")
    print(f"📦 Sản phẩm: {SanPham.objects.count()}")
    print(f"🖼️ Ảnh sản phẩm: {SanPhamImage.objects.count()}")
    print(f"🧾 Hóa đơn: {HoaDon.objects.count()}")
    print("🔐 Admin: admin@bhx.vn / admin123")
    print("🔐 Mật khẩu mặc định STAFF: 123456")
    print("💰 Lương mỗi STAFF: 2.000.000")
    print("✅ 40 order đều ở trạng thái HOAN_TAT")

    if created_customers:
        print("\nDanh sách tài khoản khách hàng trong Excel:")
        for kh in created_customers:
            if kh.tai_khoan:
                print(f"- {kh.ten_khach_hang} | {kh.dien_thoai} | {kh.tai_khoan.email}")


 
# 8. MAIN
 
if __name__ == "__main__":
    run()
