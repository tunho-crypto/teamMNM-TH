from django.contrib.auth.hashers import check_password, make_password
import random
from datetime import timedelta
import pandas as pd
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.utils import timezone
from django.core.mail import send_mail
from django.utils.html import strip_tags
import openpyxl
from openpyxl.styles import PatternFill
from django.http import HttpResponse
from .models import EmailOTP
from .models import (
    SanPham, LoaiSanPham, ChiNhanh, TaiKhoan, 
    KhachHang, NhanVien, HoaDon, ChiTietHoaDon, BinhLuanSanPham
)
from .serializers import (
    SanPhamSerializer, LoaiSanPhamSerializer, 
    ChiNhanhSerializer, BinhLuanSanPhamSerializer
)

# ==========================================
# 1. NHÓM API SẢN PHẨM & DANH MỤC
# ==========================================

@api_view(['GET'])
def get_danh_sach_san_pham(request):
    """API xuất danh sách sản phẩm sang cho VueJS lấy về hiển thị"""
    san_pham = SanPham.objects.all()
    serializer = SanPhamSerializer(san_pham, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_danh_sach_loai_san_pham(request):
    """API xuất tất cả danh mục hàng hóa (LoaiSanPham) sang VueJS"""
    loai_sp = LoaiSanPham.objects.all()
    serializer = LoaiSanPhamSerializer(loai_sp, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def get_chi_tiet_san_pham(request, pk):
    """API lấy toàn bộ thông tin cấu trúc chi tiết của 1 sản phẩm theo ID"""
    try:
        san_pham = SanPham.objects.get(pk=pk)
        serializer = SanPhamSerializer(san_pham)
        return Response(serializer.data)
    except SanPham.DoesNotExist:
        return Response({"error": "Không tìm thấy sản phẩm này sếp ơi!"}, status=404)


# ==========================================
# 2. NHÓM API BẢN ĐỒ & CHI NHÁNH (GIS)
# ==========================================

@api_view(['GET'])
def get_danh_sach_chi_nhanh(request):
    """API xuất danh sách chi nhánh và tọa độ không gian sang bản đồ VueJS"""
    chi_nhanh = ChiNhanh.objects.all()
    serializer = ChiNhanhSerializer(chi_nhanh, many=True)
    return Response(serializer.data)


# ==========================================
# 3. NHÓM API AUTHENTICATION (ĐANG NHẬP / ĐANG KÝ / OTP)
# ==========================================

@api_view(['POST'])
def dang_nhap_api(request):
    data = request.data
    username = data.get('username')
    password = data.get('password')

    tai_khoan = TaiKhoan.objects.filter(email=username).first() or TaiKhoan.objects.filter(dien_thoai=username).first()

    # SỬA Ở ĐÂY: Dùng check_password để giải mã và so sánh khớp với hệ thống cũ của sếp
    if tai_khoan and check_password(password, tai_khoan.password_hash):
        if tai_khoan.trang_thai == 'LOCKED':
            return Response({"error": "Tài khoản của sếp đã bị khóa!"}, status=400)
            
        khach_hang = getattr(tai_khoan, 'khach_hang', None)
        return Response({
            "success": True,
            "user": {
                "id": tai_khoan.id,
                "ho_ten": tai_khoan.ho_ten,
                "role": tai_khoan.role,
                "khach_hang_id": khach_hang.id if khach_hang else None
            }
        })
    return Response({"error": "Tài khoản hoặc mật khẩu không chính xác!"}, status=400)


# ==========================================
# API CHO CÁC TRANG QUẢN TRỊ ADMIN (CRUD)
# ==========================================
@api_view(['GET'])
def admin_get_accounts(request):
    """Lấy danh sách Tài khoản"""
    q = request.GET.get('q', '').strip()
    accounts = TaiKhoan.objects.all().order_by('-id')
    if q:
        from django.db.models import Q
        accounts = accounts.filter(Q(ho_ten__icontains=q) | Q(email__icontains=q) | Q(dien_thoai__icontains=q))
        
    data = [{"id": a.id, "ho_ten": a.ho_ten, "email": a.email, "dien_thoai": a.dien_thoai, "role": a.role, "trang_thai": a.trang_thai} for a in accounts]
    return Response(data)

@api_view(['GET'])
def admin_get_branches(request):
    """Lấy danh sách Chi nhánh"""
    q = request.GET.get('q', '').strip()
    branches = ChiNhanh.objects.all().order_by('-id')
    if q:
        branches = branches.filter(ten_chi_nhanh__icontains=q)
        
    data = [{"id": b.id, "ten_chi_nhanh": b.ten_chi_nhanh, "dia_chi": b.dia_chi, "dien_thoai": b.dien_thoai} for b in branches]
    return Response(data)

@api_view(['GET'])
def admin_get_customers(request):
    """Lấy danh sách Khách hàng"""
    q = request.GET.get('q', '').strip()
    customers = KhachHang.objects.all().order_by('-id')
    if q:
        customers = customers.filter(ten_khach_hang__icontains=q)
        
    data = [{"id": c.id, "ten_khach_hang": c.ten_khach_hang, "email": c.email, "dien_thoai": c.dien_thoai} for c in customers]
    return Response(data)
@api_view(['POST'])
def dang_ky_api(request):
    """API Đăng ký tài khoản khách hàng mới"""
    data = request.data
    ho_ten = data.get('ho_ten')
    email = data.get('email')
    dien_thoai = data.get('dien_thoai')
    password = data.get('password')

    if TaiKhoan.objects.filter(email=email).exists() or TaiKhoan.objects.filter(dien_thoai=dien_thoai).exists():
        return Response({"error": "Email hoặc Số điện thoại này đã được đăng ký!"}, status=400)

    tai_khoan = TaiKhoan.objects.create(
        ho_ten=ho_ten,
        email=email,
        dien_thoai=dien_thoai,
        password_hash=password,
        role='USER',
        trang_thai='ACTIVE'
    )

    KhachHang.objects.create(
        ten_khach_hang=ho_ten,
        dien_thoai=dien_thoai,
        email=email,
        tai_khoan=tai_khoan
    )
    return Response({"success": True, "message": "Đăng ký tài khoản thành công rồi sếp!"})


# ==========================================
# 4. NHÓM API THANH TOÁN & ĐƠN HÀNG
# ==========================================

@api_view(['POST'])
def xu_ly_thanh_toan_api(request):
    """API Tiếp nhận đơn hàng nâng cao, hỗ trợ hình thức và gửi mail hóa đơn HTML"""
    data = request.data
    khach_hang_id = data.get('khach_hang_id')
    
    khach_hang = KhachHang.objects.filter(id=khach_hang_id).first() if khach_hang_id else KhachHang.objects.first()
    
    delivery_method = data.get('delivery_method', 'delivery')
    payment_method = data.get('payment_method', 'BANK')
    time_slot = data.get('delivery_time_slot', 'Trong ngày')
    dia_chi = data.get('dia_chi_giao_hang', 'Địa chỉ mặc định')
    
    if delivery_method == 'pickup':
        dia_chi = f"Nhận trực tiếp tại siêu thị chi nhánh"

    hoa_don = HoaDon.objects.create(
        ngay_lap=timezone.now(),
        khach_hang=khach_hang,
        nhan_vien_id=1,
        ten_nguoi_nhan=data.get('ten_nguoi_nhan', 'Khách vãng lai'),
        sdt_nguoi_nhan=data.get('sdt_nguoi_nhan', '0123456789'),
        dia_chi_giao_hang=dia_chi,
        trang_thai='CHO_XU_LY',
        phuong_thuc_thanh_toan=payment_method
    )

    items = data.get('items', [])
    tong_tien = 0
    for item in items:
        san_pham = SanPham.objects.get(id=item['id'])
        so_luong = int(item.get('so_luong', 1))
        thanh_tien = float(san_pham.don_gia * so_luong)
        tong_tien += thanh_tien
        
        ChiTietHoaDon.objects.create(
            hoa_don=hoa_don,
            san_pham=san_pham,
            so_luong=so_luong,
            don_gia=san_pham.don_gia
        )
        if san_pham.so_luong >= so_luong:
            san_pham.so_luong -= so_luong
            san_pham.save()

    # Gửi mail HTML cho khách (Đồng bộ logic từ file cart.py của sếp)
    try:
        tai_khoan = getattr(khach_hang, 'tai_khoan', None) if khach_hang else None
        email_khach = getattr(tai_khoan, 'email', None) if tai_khoan else None
        if email_khach:
            subject = f'Xác nhận đơn hàng #{hoa_don.id} từ Bách Hóa Xanh'
            html_message = f"""
            <div style="font-family: Arial; max-width: 600px; margin: auto; border: 1px solid #ddd; padding: 20px; border-radius: 8px;">
                <h2 style="color: #008a37; text-align: center;">CẢM ƠN SẾP ĐẶT HÀNG! 🎉</h2>
                <p>Đơn hàng <strong>#{hoa_don.id}</strong> đã được lưu thành công.</p>
                <p>⏰ Khung giờ nhận: <strong>{time_slot}</strong></p>
                <p>💰 Tổng tiền: <strong style="color: red;">{tong_tien:,.0f} VNĐ</strong></p>
            </div>
            """
            send_mail(subject, strip_tags(html_message), 'no-reply@bhx.local', [email_khach], html_message=html_message)
    except Exception as e:
        print(f"Lỗi gửi mail: {e}")

    return Response({"success": True, "message": f"Đã lưu hóa đơn #{hoa_don.id} và gửi mail xác nhận!"})

@api_view(['GET'])
def lich_su_don_hang_api(request, khach_hang_id):
    """API lấy lịch sử đơn đặt hàng của một khách hàng cụ thể"""
    don_hangs = HoaDon.objects.filter(khach_hang_id=khach_hang_id).order_by('-ngay_lap')
    data = []
    for dh in don_hangs:
        chi_tiets = ChiTietHoaDon.objects.filter(hoa_don=dh)
        items_data = [{
            "ten_san_pham": ct.san_pham.ten_san_pham,
            "so_luong": ct.so_luong,
            "don_gia": float(ct.don_gia),
            "thanh_tien": float(ct.so_luong * ct.don_gia)
        } for ct in chi_tiets]
        
        data.append({
            "id": dh.id,
            "ngay_lap": dh.ngay_lap.strftime("%d/%m/%Y %H:%M"),
            "trang_thai": dh.trang_thai,
            "trang_thai_hien_thi": dh.get_trang_thai_display(),
            "tong_tien": sum(item["thanh_tien"] for item in items_data),
            "dia_chi_giao_hang": dh.dia_chi_giao_hang,
            "items": items_data
        })
    return Response(data)

@api_view(['POST'])
def hu_don_hang_api(request, order_id):
    """API hủy đơn hàng ở trạng thái chờ duyệt"""
    try:
        hoa_don = HoaDon.objects.get(id=order_id)
        if hoa_don.trang_thai == 'CHO_XU_LY':
            hoa_don.trang_thai = 'DA_HUY'
            hoa_don.save(update_fields=['trang_thai'])
            return Response({"success": True, "message": f"Đã hủy đơn hàng #{order_id} thành công!"})
        return Response({"error": "Đơn hàng này không thể hủy."}, status=400)
    except HoaDon.DoesNotExist:
        return Response({"error": "Không tìm thấy hóa đơn."}, status=404)


# ==========================================
# 5. NHÓM API ĐÁNH GIÁ & BÌNH LUẬN
# ==========================================

@api_view(['GET'])
def get_binh_luan_san_pham(request, pk):
    """API xuất danh sách bình luận của sản phẩm"""
    binh_luan = BinhLuanSanPham.objects.filter(san_pham_id=pk).order_by('-ngay_binh_luan')
    serializer = BinhLuanSanPhamSerializer(binh_luan, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def gui_binh_luan_api(request, pk):
    """API lưu bình luận mới vào cơ sở dữ liệu"""
    data = request.data
    tai_khoan_id = data.get('tai_khoan_id')
    noi_dung = data.get('noi_dung')
    so_sao = data.get('so_sao', 5)

    if not noi_dung:
        return Response({"error": "Nội dung trống!"}, status=400)
    try:
        BinhLuanSanPham.objects.create(
            san_pham_id=pk,
            tai_khoan_id=tai_khoan_id,
            noi_dung=noi_dung,
            so_sao=int(so_sao),
            ngay_binh_luan=timezone.now()
        )
        return Response({"success": True, "message": "Đăng đánh giá thành công!"})
    except Exception:
        return Response({"error": "Lỗi hệ thống!"}, status=400)

@api_view(['POST'])
def edit_product_comment_api(request, pk):
    try:
        binh_luan = BinhLuanSanPham.objects.get(id=pk)
        binh_luan.noi_dung = request.data.get('noi_dung', binh_luan.noi_dung)
        binh_luan.save()
        return Response({"success": True, "message": "Cập nhật thành công!"})
    except BinhLuanSanPham.DoesNotExist:
        return Response({"error": "Không tìm thấy"}, status=404)

@api_view(['DELETE'])
def delete_product_comment_api(request, pk):
    try:
        binh_luan = BinhLuanSanPham.objects.get(id=pk)
        binh_luan.delete()
        return Response({"success": True, "message": "Đã xóa!"})
    except BinhLuanSanPham.DoesNotExist:
        return Response({"error": "Không tìm thấy"}, status=404)


# ==========================================
# 6. NHÓM API QUÂN TRỊ EXCEL & DASHBOARD
# ==========================================

@api_view(['GET'])
def admin_dashboard_stats_api(request):
    don_hang_thanh_cong = HoaDon.objects.exclude(trang_thai__in=['CHO_XU_LY', 'DA_HUY'])
    tong_doanh_thu = 0
    for dh in don_hang_thanh_cong:
        tong_doanh_thu += sum(ct.so_luong * ct.don_gia for ct in ChiTietHoaDon.objects.filter(hoa_don=dh))

    return Response({
        "tong_san_pham": SanPham.objects.count(),
        "tong_don_hang": HoaDon.objects.count(),
        "tong_doanh_thu": tong_doanh_thu,
        "san_pham_sap_het": [{
            "id": sp.id, "ten_san_pham": sp.ten_san_pham, "so_luong": sp.so_luong
        } for sp in SanPham.objects.filter(so_luong__lte=5).order_by('so_luong')[:5]]
    })

@api_view(['GET'])
def export_excel_api(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doanh thu"
    header_fill = PatternFill(start_color="008A37", end_color="008A37", fill_type="solid")
    ws.append(['Mã Đơn', 'Ngày Lập', 'Địa Chỉ Giao', 'Trạng Thái', 'Tổng Tiền (VNĐ)'])
    for cell in ws[1]:
        cell.fill = header_fill

    thang_nam = request.GET.get('thang_nam')
    orders_qs = HoaDon.objects.all().order_by('-ngay_lap')
    if thang_nam:
        nam, thang = thang_nam.split('-')
        orders_qs = orders_qs.filter(ngay_lap__year=nam, ngay_lap__month=thang)

    for order in orders_qs:
        total = sum(item.so_luong * item.don_gia for item in ChiTietHoaDon.objects.filter(hoa_don=order))
        ws.append([
            order.id,
            order.ngay_lap.strftime("%d/%m/%Y %H:%M"),
            order.dia_chi_giao_hang,
            order.get_trang_thai_display(),
            total
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ten_file = f'Bao_Cao_Doanh_Thu_{thang_nam}.xlsx' if thang_nam else 'Bao_Cao_Doanh_Thu_Tat_Ca.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{ten_file}"'
    wb.save(response)
    return response

@api_view(['POST'])
def import_excel_api(request):
    file_excel = request.FILES.get('file_excel')
    if not file_excel: return Response({"error": "Thiếu file!"}, status=400)
    try:
        df = pd.read_excel(file_excel)
        for _, row in df.iterrows():
            loai_obj, _ = LoaiSanPham.objects.get_or_create(ten_loai=str(row.get('ten_loai', 'Khác')))
            SanPham.objects.update_or_create(
                ten_san_pham=row['ten_san_pham'],
                defaults={'loai': loai_obj, 'don_gia': row.get('don_gia', 0), 'so_luong': row.get('so_luong', 0), 'mo_ta': row.get('mo_ta', ''), 'gia_hien_tai': row.get('don_gia', 0)}
            )
        return Response({"success": True, "message": "Nạp file Excel thành công!"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)

from django.db.models import Q
from .models import ChiNhanh, BinhLuanChiNhanh, BinhLuanChiNhanhImage, TonKhoChiNhanh, NhanVien

# --- 1. API PHÂN HỆ CỬA HÀNG & ĐÁNH GIÁ ĐA PHƯƠNG TIỆN ---
@api_view(['GET'])
def api_danh_sach_cua_hang(request):
    """API tìm kiếm cửa hàng và lọc động theo Quận/Huyện"""
    quan_huyen = request.GET.get('quan')
    tu_khoa = request.GET.get('q')
    stores = ChiNhanh.objects.all()

    if quan_huyen and quan_huyen != 'all':
        stores = stores.filter(dia_chi__icontains=quan_huyen)
    if tu_khoa:
        stores = stores.filter(Q(ten_chi_nhanh__icontains=tu_khoa) | Q(dia_chi__icontains=tu_khoa))

    data = [{
        "id": st.id, "ten_chi_nhanh": st.ten_chi_nhanh,
        "dia_chi": st.dia_chi, "dien_thoai": st.dien_thoai
    } for st in stores]
    return Response(data)

@api_view(['GET'])
def api_chi_tiet_cua_hang(request, pk):
    """API lấy thông tin chi nhánh và danh sách bình luận kèm nhiều ảnh của khách"""
    try:
        store = ChiNhanh.objects.get(pk=pk)
        comments = BinhLuanChiNhanh.objects.filter(chi_nhanh=store).select_related('tai_khoan').order_by("-created_at")
        
        comments_data = []
        for cm in comments:
            # Quét mớ ảnh phụ đi kèm bình luận chi nhánh
            imgs = BinhLuanChiNhanhImage.objects.filter(binh_luan=cm)
            img_urls = [request.build_absolute_uri(i.image.url) for i in imgs if i.image]
            
            comments_data.append({
                "id": cm.id,
                "ten_nguoi_dung": cm.tai_khoan.ho_ten,
                "noi_dung": cm.noi_dung, "so_sao": cm.so_sao,
                "hinh_anhs": img_urls,
                "created_at": cm.created_at.strftime("%d/%m/%Y %H:%M")
            })
            
        return Response({
            "id": store.id, "ten_chi_nhanh": store.ten_chi_nhanh,
            "dia_chi": store.dia_chi, "dien_thoai": store.dien_thoai,
            "comments": comments_data
        })
    except ChiNhanh.DoesNotExist:
        return Response({"error": "Không tìm thấy chi nhánh"}, status=404)

@api_view(['POST'])
def api_gui_binh_luan_cua_hang(request, pk):
    """API tiếp nhận đánh giá cửa hàng và xử lý lưu nhiều ảnh phụ (Multi-image Upload)"""
    tai_khoan_id = request.data.get('tai_khoan_id')
    noi_dung = request.data.get('noi_dung')
    so_sao = int(request.data.get('so_sao', 5))

    if not tai_khoan_id or not noi_dung:
        return Response({"error": "Dữ liệu không đầy đủ sếp ơi!"}, status=400)

    # Khởi tạo bản ghi đánh giá chi nhánh
    binh_luan = BinhLuanChiNhanh.objects.create(
        chi_nhanh_id=pk, tai_khoan_id=tai_khoan_id,
        noi_dung=noi_dung, so_sao=so_sao
    )

    # Đọc mớ tệp đa phương tiện gửi lên từ VueJS dán thẳng vào bảng chứa ảnh
    images = request.FILES.getlist('hinh_anh_danh_gia')
    for img in images:
        BinhLuanChiNhanhImage.objects.create(binh_luan=binh_luan, image=img)

    return Response({"success": True, "message": "Đã ghi nhận đánh giá đa phương tiện thành công!"})


# --- 2. API PHÂN HỆ QUẢN LÝ TỒN KHO PHÂN QUYỀN (ADMIN & STAFF) ---
@api_view(['GET'])
def api_quan_ly_ton_kho(request):
    """API lấy dữ liệu tồn kho động. Tự nhận diện quyền Admin/Staff và lọc trạng thái"""
    role = request.GET.get('role', 'USER')
    user_id = request.GET.get('user_id')
    keyword = request.GET.get('q', '').strip()
    trang_thai = request.GET.get('trang_thai', '').strip()

    ton_khos = TonKhoChiNhanh.objects.select_related("chi_nhanh", "san_pham", "san_pham__loai").all()

    # Khóa logic phân quyền: Nếu là STAFF thì ép chỉ xem kho của chi nhánh mình phụ trách
    if role == 'STAFF' and user_id:
        nv = NhanVien.objects.filter(tai_khoan_id=user_id).first()
        if nv and nv.chi_nhanh_id:
            ton_khos = ton_khos.filter(chi_nhanh=nv.chi_nhanh)
        else:
            return Response([])

    if keyword:
        ton_khos = ton_khos.filter(Q(san_pham__ten_san_pham__icontains=keyword) | Q(san_pham__loai__ten_loai__icontains=keyword))

    # Bộ lọc trạng thái thông minh từ file employees.py
    if trang_thai == "HET_HANG":
        ton_khos = ton_khos.filter(so_luong_ton__lte=0)
    elif trang_thai == "SAP_HET":
        ton_khos = ton_khos.filter(so_luong_ton__gt=0, so_luong_ton__lte=5)
    elif trang_thai == "CON_HANG":
        ton_khos = ton_khos.filter(so_luong_ton__gt=5)

    data = [{
        "id": tk.id, "ten_chi_nhanh": tk.chi_nhanh.ten_chi_nhanh,
        "ten_san_pham": tk.san_pham.ten_san_pham, "ten_loai": tk.san_pham.loai.ten_loai,
        "so_luong_ton": tk.so_luong_ton, "muc_can_canh_bao": tk.muc_can_canh_bao
    } for tk in ton_khos]
    return Response(data)

@api_view(['POST'])
def api_cap_nhat_ton_kho(request, pk):
    """API cập nhật trực tiếp số lượng hoặc mức cảnh báo tồn kho chi nhánh"""
    try:
        ton_kho = TonKhoChiNhanh.objects.get(pk=pk)
        action_type = request.data.get('action_type', 'set') # 'set' hoặc 'add'
        
        if action_type == 'add':
            ton_kho.so_luong_ton += int(request.data.get('so_luong_nhap_them', 0))
        else:
            ton_kho.so_luong_ton = int(request.data.get('so_luong_ton', ton_kho.so_luong_ton))
            
        ton_kho.muc_can_canh_bao = int(request.data.get('muc_can_canh_bao', ton_kho.muc_can_canh_bao))
        ton_kho.save()
        return Response({"success": True, "message": "Cập nhật kho thành công!"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)
    
# ==========================================
# NHÓM API XÁC THỰC OTP & QUÊN MẬT KHẨU
# ==========================================

@api_view(['POST'])
def xac_thuc_otp_api(request):
    """API kiểm tra mã OTP kích hoạt tài khoản"""
    email = request.data.get('email')
    otp_code = request.data.get('otp_code')
    
    # Tìm mã OTP còn hạn và chưa sử dụng
    otp_record = EmailOTP.objects.filter(
        email=email, 
        otp_code=otp_code, 
        is_verified=False,
        expires_at__gt=timezone.now()
    ).first()
    
    if otp_record:
        otp_record.is_verified = True
        otp_record.save()
        # Mở khóa tài khoản thành ACTIVE
        TaiKhoan.objects.filter(email=email).update(trang_thai='ACTIVE')
        return Response({"success": True, "message": "Kích hoạt tài khoản thành công!"})
        
    return Response({"error": "Mã OTP không chính xác hoặc đã hết hạn!"}, status=400)

@api_view(['POST'])
def gui_lai_otp_api(request):
    """API sinh mã OTP mới và gửi lại qua Email"""
    email = request.data.get('email')
    if not TaiKhoan.objects.filter(email=email).exists():
        return Response({"error": "Email không tồn tại trong hệ thống!"}, status=400)
        
    otp_code = str(random.randint(100000, 999999))
    EmailOTP.objects.create(
        email=email,
        otp_code=otp_code,
        expires_at=timezone.now() + timedelta(minutes=5)
    )
    
    send_mail(
        'Mã xác thực OTP Bách Hóa Xanh',
        f'Mã xác thực tài khoản của sếp là: {otp_code}. Mã có hiệu lực trong 5 phút.',
        'no-reply@bhx.local',
        [email]
    )
    return Response({"success": True, "message": "Đã gửi lại mã OTP!"})

@api_view(['POST'])
def quen_mat_khau_api(request):
    """API gửi OTP khôi phục mật khẩu"""
    email = request.data.get('email')
    tai_khoan = TaiKhoan.objects.filter(email=email).first()
    if not tai_khoan:
        return Response({"error": "Email này chưa được đăng ký!"}, status=400)
        
    otp_code = str(random.randint(100000, 999999))
    EmailOTP.objects.create(
        email=email,
        tai_khoan=tai_khoan,  # Bắt buộc phải có - fix lỗi NOT NULL
        otp_code=otp_code,
        expires_at=timezone.now() + timedelta(minutes=15)
    )
    send_mail(
        'Khôi phục mật khẩu Bách Hóa Xanh',
        f'Mã OTP khôi phục mật khẩu của sếp là: {otp_code}',
        'no-reply@bhx.local',
        [email]
    )
    return Response({"success": True, "message": "Đã gửi mã khôi phục!"})

@api_view(['POST'])
def dat_lai_mat_khau_api(request):
    """API cập nhật mật khẩu mới (có mã hóa Hash)"""
    otp_code = request.data.get('otp_code')
    new_password = request.data.get('new_password')
    
    otp_record = EmailOTP.objects.filter(
        otp_code=otp_code,
        is_verified=False,
        expires_at__gt=timezone.now()
    ).first()
    
    if otp_record:
        tai_khoan = TaiKhoan.objects.filter(email=otp_record.email).first()
        if tai_khoan:
            tai_khoan.password_hash = make_password(new_password)
            tai_khoan.save()
            
            otp_record.is_verified = True
            otp_record.save()
            return Response({"success": True, "message": "Đổi mật khẩu thành công!"})
            
    return Response({"error": "Mã OTP không hợp lệ hoặc đã quá hạn!"}, status=400)


# ==========================================
# NHÓM API THANH TOÁN TẠI QUẦY (POS)
# ==========================================

@api_view(['POST'])
def pos_thanh_toan_api(request):
    """API xử lý đơn hàng quét mã trực tiếp tại quầy thu ngân"""
    data = request.data
    tien_khach_dua = float(data.get('tien_khach_dua', 0))
    items = data.get('items', [])
    
    if not items:
        return Response({"error": "Đơn hàng đang trống!"}, status=400)
        
    # Tạo hóa đơn HOÀN TẤT ngay lập tức vì mua trực tiếp tại quầy
    hoa_don = HoaDon.objects.create(
        ngay_lap=timezone.now(),
        ten_nguoi_nhan="Khách mua trực tiếp",
        dia_chi_giao_hang="Tại quầy",
        trang_thai="HOAN_TAT",
        phuong_thuc_thanh_toan="CASH",
        nhan_vien_id=1 # Giả lập ID của nhân viên đang trực máy POS
    )
    
    tong_tien = 0
    invoice_items = []
    
    for item in items:
        sp = SanPham.objects.get(id=item['id'])
        sl = int(item.get('so_luong_mua', 1))
        thanh_tien = sl * float(sp.don_gia)
        tong_tien += thanh_tien
        
        ChiTietHoaDon.objects.create(
            hoa_don=hoa_don,
            san_pham=sp,
            so_luong=sl,
            don_gia=sp.don_gia
        )
        
        # Cập nhật tồn kho tự động
        if sp.so_luong >= sl:
            sp.so_luong -= sl
            sp.save()
            
        # Đóng gói dữ liệu để VueJS xuất file in nhiệt
        invoice_items.append({
            "ten_san_pham": sp.ten_san_pham,
            "so_luong_mua": sl,
            "don_gia": float(sp.don_gia),
            "thanh_tien": thanh_tien
        })
        
    tien_hoan_lai = tien_khach_dua - tong_tien
    
    invoice_data = {
        "hoa_don_id": hoa_don.id,
        "ngay_gio": hoa_don.ngay_lap.strftime("%d/%m/%Y %H:%M:%S"),
        "thu_ngan": "Nhân viên quầy số 1",
        "tong_tien": tong_tien,
        "tien_khach_dua": tien_khach_dua,
        "tien_hoan_lai": tien_hoan_lai if tien_hoan_lai > 0 else 0,
        "items": invoice_items
    }
    
    return Response({"success": True, "invoice_data": invoice_data})
# ==========================================
# BỌC THÉP API XUẤT DANH SÁCH SẢN PHẨM ADMIN
# ==========================================
@api_view(['GET'])
def admin_get_products(request):
    q = request.GET.get('q', '').strip()
    loai = request.GET.get('loai', '')
    products = SanPham.objects.all().order_by('-id')
    
    if q: products = products.filter(ten_san_pham__icontains=q)
    if loai: products = products.filter(loai_id=loai)
        
    import os
    data = []
    for p in products:
        # 1. Lay anh qua anh_chinh (related object) dung nhu Django template
        img_url = None
        try:
            anh_chinh = p.anh_chinh
            if anh_chinh and anh_chinh.image and anh_chinh.image.name:
                if os.path.isfile(anh_chinh.image.path):
                    img_url = request.build_absolute_uri(anh_chinh.image.url)
        except Exception:
            img_url = None

        # 2. Xu ly Loai san pham
        loai_ten = "Chua phan loai"
        try:
            if p.loai:
                loai_ten = p.loai.ten_loai
        except Exception:
            pass

        # 3. Xu ly gia (uu tien gia khuyen mai nhu Django template)
        is_khuyen_mai = getattr(p, 'is_khuyen_mai', False)
        gia_khuyen_mai = getattr(p, 'gia_khuyen_mai', None)
        don_gia = float(p.don_gia) if getattr(p, 'don_gia', None) else 0
        gia_hien_thi = float(gia_khuyen_mai) if (is_khuyen_mai and gia_khuyen_mai) else don_gia

        data.append({
            "id": p.id,
            "ten_san_pham": p.ten_san_pham,
            "loai_ten": loai_ten,
            "don_gia": don_gia,
            "gia_hien_thi": gia_hien_thi,
            "is_khuyen_mai": is_khuyen_mai,
            "so_luong": getattr(p, 'so_luong', 0),
            "hinh_anh_url": img_url
        })
    return Response(data)

@api_view(['GET'])
def admin_get_customers(request):
    q = request.GET.get('q', '').strip()
    customers = KhachHang.objects.all().order_by('-id')
    if q:
        customers = customers.filter(ten_khach_hang__icontains=q)
        
    data = []
    for c in customers:
        # Sửa lại: Lấy dia_chi thay vì email theo đúng thiết kế CSDL của sếp
        data.append({
            "id": c.id, 
            "ten_khach_hang": c.ten_khach_hang, 
            "dien_thoai": getattr(c, 'dien_thoai', ''),
            "dia_chi": getattr(c, 'dia_chi', '')
        })
    return Response(data)

@api_view(['GET'])
def admin_get_orders(request):
    """API xuất danh sách hóa đơn cho bảng Admin"""
    q = request.GET.get('q', '')
    trang_thai = request.GET.get('trang_thai', '')
    orders = HoaDon.objects.all().order_by('-ngay_lap')
    
    if q: orders = orders.filter(ten_nguoi_nhan__icontains=q)
    if trang_thai: orders = orders.filter(trang_thai=trang_thai)
        
    data = [{
        "id": o.id, 
        "ten_nguoi_nhan": o.ten_nguoi_nhan, 
        "sdt_nguoi_nhan": o.sdt_nguoi_nhan,
        "ngay_lap": o.ngay_lap.strftime("%d/%m/%Y %H:%M"), 
        "trang_thai": o.get_trang_thai_display()
    } for o in orders]
    return Response(data)
    
@api_view(['POST'])
def admin_update_order_status(request, pk):
    """API duyệt / thay đổi trạng thái hóa đơn"""
    try:
        order = HoaDon.objects.get(pk=pk)
        order.trang_thai = request.data.get('trang_thai', order.trang_thai)
        order.save(update_fields=['trang_thai'])
        return Response({"success": True})
    except HoaDon.DoesNotExist:
        return Response({"error": "Không tìm thấy đơn hàng!"}, status=404)